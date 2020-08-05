from tkinter import *
from openpyxl import *


# globally declare wb and sheet variable

# opening the existing excel file

wb = load_workbook('sample.xlsx')
# create the sheet object 
sheet = wb.active


def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30  # Name
    sheet.column_dimensions['B'].width = 30  # Caste
    sheet.column_dimensions['C'].width = 30  # Star
    sheet.column_dimensions['D'].width = 30  # Phone
    sheet.column_dimensions['D'].width = 50  # Address
    # sheet.column_dimensions['D'].width = 30

    # write given data to an excel spreadsheet
    # at particular location

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Caste"
    sheet.cell(row=1, column=3).value = "Star"
    sheet.cell(row=1, column=4).value = "Phone"
    sheet.cell(row=1, column=5).value = "Address"


# Function to set focus (cursor)
def focus1(event):
    # Function to set focus Name
    Name_field.focus_set()


# Function to set focus 
def focus2(event):
    # set focus on the Caste
    Caste_field.focus_set()


# Function to set focus 
def focus3(event):
    # set focus on the form Star 
    Star_field.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the form Phone 
    Phone_field.focus_set()


# Function to set focus
def focus5(event):
    # set focus on the form Address 
    Address_field.focus_set()


# Function for clearing the
# contents of text entry boxes 
def clear():
    # clear the content of text entry box 
    Name_field.delete(0, END)
    Caste_field.delete(0, END)
    Phone_field.delete(0, END)
    Address_field.delete(0, END)
    Star_field.delete(0, END)


# Function to take data from GUI  
# window and write to an excel file 
def insert():
    # if user not fill any entry 
    # then print "empty input" 
    if (Name_field.get() == "" and
            Caste_field.get() == "" and
            Phone_field.get() == "" and
            Star_field.get() == "" and
            Address_field.get() == ""):

        print("empty input")

    else:

        # assigning the max row and max column 
        # value upto which data is written 
        # in an excel sheet to the variable 
        current_row = sheet.max_row
        current_column = sheet.max_column

        # get method returns current text 
        # as string which we write into 
        # excel spreadsheet at particular location 
        sheet.cell(row=current_row + 1, column=1).value = Name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = Caste_field.get()
        sheet.cell(row=current_row + 1, column=3).value = Phone_field.get()
        sheet.cell(row=current_row + 1, column=4).value = Star_field.get()
        sheet.cell(row=current_row + 1, column=5).value = Address_field.get()

        # save the file
        wb.save('C:\\Users\\user\\Documents\\Malayogam.xlsx')

        # set focus on the name_field box 
        Name_field.focus_set()

        # call the clear() function 
        clear()

        # driven code


if __name__ == "__main_":
    # create a GUI window
    root = Tk()

    # set the background colour of GUI window

    root.configure(background='lightgreen')

    # set the title of the window

    root.title("Malayogam Form")

    # set the configuration of GUI window

    root.geometry("500x800")

    excel()

    # create a form label

    heading = Label(root, text="Form", bg="lightgreen")

    # create a Name Label

    Name = Label(root, text="Name : ", bg="lightgreen")

    # create a caste field

    Caste = Label(root, text="Caste : ", bg="lightgreen")

    # create a Phone

    Phone = Label(root, text="phone : ", bg="lightgreen")

    # create a star

    Star = Label(root, text="Star : ", bg="lightgreen")

    # create an address label

    Address = Label(root, text="Address : ", bg="lightgreen")

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    Name.grid(row=1, column=0)
    Caste.grid(row=2, column=0)
    Phone.grid(row=3, column=0)
    Star.grid(row=4, column=0)
    Address.grid(row=5, column=0)

    # create a text entry box
    # for typing the information
    Name_field = Entry(root)
    Caste_field = Entry(root)
    Phone_field = Entry(root)
    Star_field = Entry(root)
    Address_field = Entry(root)

    # bind method of widget is used for
    # the binding the function with the events

    # whenever the enter key is pressed
    # then call the focus1 function
    Name_field.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus2 function
    Caste_field.bind("<Return>", focus2)

    # whenever the enter key is pressed
    # then call the focus3 function
    Phone_field.bind("<Return>", focus3)

    # whenever the enter key is pressed
    # then call the focus4 function
    Star_field.bind("<Return>", focus4)

    # whenever the enter key is pressed
    # then call the focus5 function
    Address_field.bind("<Return>", focus5)

    # whenever the enter key is pressed
    # then call the focus6 function

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    Name_field.grid(row=1, column=1, ipadx="100")
    Caste_field.grid(row=2, column=1, ipadx="100")
    Phone_field.grid(row=3, column=1, ipadx="100")
    Star_field.grid(row=4, column=1, ipadx="100")
    Address_field.grid(row=5, column=1, ipadx="100")

    # call excel function
    excel()

    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black",
                    bg="Red", command=insert)
    submit.grid(row=8, column=1)

    # start the GUI
mainloop()
